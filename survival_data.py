import glob 
import numpy as np
import os
from datetime import date
import openpyxl
import csv
import math 
import tensorflow_addons as tfa
from tensorflow.keras.callbacks import ReduceLROnPlateau, ModelCheckpoint, EarlyStopping, TensorBoard
from sklearn.model_selection import train_test_split
from datetime import datetime
from experiments.exp_3d.preprocessing import *
from dataGenerator import *
from VnetSurvival import *

""" https://nbviewer.jupyter.org/github/sebp/survival-cnn-estimator/blob/master/tutorial_tf2.ipynb 
    https://github.com/chl8856/DeepHit
"""

#PARAMETRES 
csv_path = ''
pp_dir = ''
modalities = ('pet_img', 'ct_img')
mode = ['binary', 'probs', 'mean_probs'][0]
method = ['otsu', 'absolute', 'relative', 'otsu_abs'][0]

#callbacks
patience = 10
ReduceLROnPlateau1 = False
EarlyStopping1 = False
ModelCheckpoint1 = False
TensorBoard1 = True

#parameters
image_shape= (256, 128, 128)
in_channels= len(modalities)
out_channels= 1
channels_last=True
keep_prob= 1.0
keep_prob_last_layer= 0.8
kernel_size= (5, 5, 5)
num_channels= 8
num_levels= 4
num_convolutions= (1, 2, 3, 3)
bottom_convolutions= 3
activation= "relu"
activation_last_layer= 'sigmoid'

architecture = 'vnet_survival'
#[if mode = binary & method = relative : t_val = 0.42
#if mode = binary & method = absolute : t_val = 2.5, 
#else : don't need tval]
tval = ''
target_size = (128, 128, 256)
target_spacing = (4.0, 4.0, 4.0)
target_direction = (1,0,0,0,1,0,0,0,1)

from_pp=False
cache_pp=False
pp_flag=''

mask=False

###Parametres dataGenerator 
epochs = 100
batch_size = 2
shuffle = True 

#FIN PARAMETRES 

def get_data_survival(base_path, excel_path, create_csv, csv_path=''):
    '''
        retrieve the survival data from the excel file and the path of the nifti files from the base path 
        create a csv file containing all the data if csv == True
    '''
    y_time=[]
    y_event=[]
    PT_paths=[]
    CT_paths=[]
    Anonym=[]
    data_exists= True #boolean path of the nifti file found
    data_dict=dict()
    #data_dict['time']=

    #open file containing survival information
    excel = openpyxl.load_workbook(excel_path)
    sheet = excel.active

    #for each patient
    for i in range(2, sheet.max_row):
        data_exists=True
        x = sheet.cell(row=i, column=5)
        if (x.value != None):
            #retrieve the path to the nifti files from the patient : folder with the Anonymisation name from the excel cell(row=i, column=2)
            path = base_path+sheet.cell(row=i, column=2).value+'/'
            nifti_path_PT= glob.glob(os.path.join(path, '**/*_nifti_PT.nii'), recursive=True)
            nifti_path_CT= glob.glob(os.path.join(path, '**/*_nifti_CT.nii'), recursive=True)
            if nifti_path_PT and nifti_path_CT:
                Anonym = np.append(Anonym,sheet.cell(row=i, column=2).value)
                PT_paths=np.append(PT_paths, [nifti_path_PT[0]])
                CT_paths=np.append(CT_paths, [nifti_path_CT[0]])
            else:
                data_exists=False

            if data_exists:
                #retrieve y_time and y_event from excel (date format (month/day/year))
                x = x.value.split('/')
                x = [int(i) for i in x]
                diagnosis_date= date(x[2],x[0],x[1])
                #if censored and there is a last check up date retrieve date
                if (sheet.cell(row=i, column=7).value==0 and sheet.cell(row=i, column=9).value!=None):
                    y= sheet.cell(row=i, column=9).value.split('/')
                    y = [int(j) for j in y]
                    last_checkup_date= date(y[2],y[0],y[1])
                #if not censored and there is a relapse date retrieve date 
                elif (sheet.cell(row=i, column=7).value==1 and sheet.cell(row=i, column=8).value!=None):
                    y= sheet.cell(row=i, column=8).value.split('/')
                    y = [int(j) for j in y]
                    last_checkup_date= date(y[2],y[0],y[1])
                
                #time is given in 30 days intervals 
                time= int(((last_checkup_date-diagnosis_date).days)/30)
                y_time=np.append(y_time, [time])
                y_event=np.append(y_event,[int(sheet.cell(row=i, column=7).value)])

    y_event= y_event.astype(np.int32)
    y_time=y_time.astype(np.int32)

    data_zip = zip(Anonym,y_time, y_event, CT_paths, PT_paths)

    if create_csv: 
        with open(csv_path, 'w') as csv_path:
            wtr = csv.writer(open (csv_path, 'w'), delimiter=',', lineterminator='\n')
            wtr.writerow(["anonymisation,time,event,CT_path,PT_path"])
            for x in data_zip : wtr.writerow ([str(x[0])+','+str(x[1])+','+str(x[2])+','+str(x[3])+','+str(x[4])])

    return list(zip(PT_paths, CT_paths)), list(zip(y_time, y_event))

trained_model_path = None #If None, train from scratch 
training_model_folder = '../model/'
now = datetime.now().strftime("%Y%m%d-%H:%M:%S")

training_model_folder = os.path.join(training_model_folder, now)  # '/path/to/folder'
if not os.path.exists(training_model_folder):
    os.makedirs(training_model_folder)
logdir = os.path.join(training_model_folder, 'logs')
if not os.path.exists(logdir):
    os.makedirs(logdir)




base_path='../../FLIP_NIFTI/'
excel_path='./FLIP.xlsx'
csv_path="CSV_FLIP.csv"
x,y= get_data_survival(base_path,excel_path,True,csv_path)
"""
#number of neurons on the output layer 
time_horizon = math.ceil(max(y)[0]*1.2)

x_train, x_val, y_train, y_val = train_test_split(x, y, test_size=0.2, random_state=42)

train_y_time=[y_t[0] for y_t in y_train]
val_y_time=[y_v[0] for y_v in y_val]
train_y_event=[y_t[1] for y_t in y_train]
val_y_event=[y_v[1] for y_v in y_val]

y_train=[train_y_time, train_y_event]
y_val=[val_y_time, val_y_event]

dataset_x = dict()
dataset_x['train']=[]
dataset_x['val']=[]

for i in range(len(x_train)):
    dataset_x['train'].append({'pet_img':x_train[i][0], 'ct_img':x_train[i][1]})
for i in range(len(x_val)):    
    dataset_x['val'].append({'pet_img':x_val[i][0], 'ct_img':x_val[i][1]})

train_transforms = get_transform('train', modalities, mask, mode, method, tval, target_size, target_spacing, target_direction, None, data_augmentation = True, from_pp=False, cache_pp=False)
val_transforms = get_transform('val', modalities, mask, mode, method, tval, target_size, target_spacing, target_direction, None,  data_augmentation = False, from_pp=False, cache_pp=False)

# multi gpu training strategy
strategy = tf.distribute.MirroredStrategy()

train_images_paths_x, val_images_paths_x = dataset_x['train'], dataset_x['val']

train_generator = DataGeneratorSurvival(train_images_paths_x,
                                        y_train,
                                        train_transforms,
                                        batch_size=batch_size,
                                        shuffle=shuffle,
                                        x_key='input')

val_generator = DataGeneratorSurvival(val_images_paths_x,
                                        y_val,
                                        val_transforms,
                                        batch_size=batch_size,
                                        shuffle=False,
                                        x_key='input')

#################################################################################
with strategy.scope():
    # definition of loss, optimizer and metrics
    loss_object = get_loss_survival(time_horizon_dim=time_horizon, batch_size=batch_size)
    optimizer = tfa.optimizers.AdamW(learning_rate=1e-4, weight_decay=1e-4)
    td_c_index = metric_td_c_index(time_horizon_dim=time_horizon, batch_size=batch_size)
    metrics = [] 

# callbacks
callbacks = []
if ReduceLROnPlateau1 == True :
    # reduces learning rate if no improvement are seen
    learning_rate_reduction = ReduceLROnPlateau(monitor= loss_object,
                                                patience=patience ,
                                                verbose=1,
                                                factor=0.5,
                                                min_lr=0.0000001)
    callbacks.append(learning_rate_reduction)

if EarlyStopping1 == True :
    # stop training if no improvements are seen
    early_stop = EarlyStopping(monitor="val_loss",
                                mode="min",
                                patience=int(patience // 2),
                                restore_best_weights=True)
    callbacks.append(early_stop)

if ModelCheckpoint1 == True :
    # saves model weights to file
    # 'model_weights.{epoch:02d}-{val_loss:.2f}.hdf5'
    checkpoint = ModelCheckpoint(os.path.join(training_model_folder, 'model_weights.h5'),
                                    monitor=loss_object,  #td_c_index ?
                                    verbose=1,
                                    save_best_only=True,
                                    mode='min',  # max
                                    save_weights_only=False)
    callbacks.append(checkpoint)

if TensorBoard1 == True :
    tensorboard_callback = TensorBoard(log_dir=logdir,
                                        histogram_freq=0,
                                        update_freq='epoch',
                                        write_graph=True,
                                        write_grads=True,
                                        write_images=True)
    callbacks.append(tensorboard_callback)

with strategy.scope():
    model = VnetSurvival(image_shape,
            in_channels,
            out_channels,
            time_horizon,
            channels_last,
            keep_prob,
            keep_prob_last_layer,
            kernel_size,
            num_channels,
            num_levels,
            num_convolutions,
            bottom_convolutions,
            activation,
            activation_last_layer).create_model()


with strategy.scope():
    model.compile(loss=loss_object, optimizer=optimizer, metrics=metrics)

if trained_model_path is not None:
    with strategy.scope():
        model.load_weights(trained_model_path)
print(model.summary())


model_json = model.to_json()
with open(os.path.join(training_model_folder, 'architecture_{}_model_{}.json'.format(architecture, now)),
            "w") as json_file:
    json_file.write(model_json)

# training model
history = model.fit(train_generator,
                    steps_per_epoch=len(train_generator),
                    validation_data=val_generator,
                    validation_steps=len(val_generator),
                    epochs=epochs,
                    callbacks=callbacks,  # initial_epoch=0,
                    verbose=1
                    )
                    """
