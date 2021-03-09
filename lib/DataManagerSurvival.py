import numpy as np
import pandas as pd
import os
import glob
import openpyxl
import csv

from datetime import date
from sklearn.model_selection import train_test_split


class DataManagerSurvival(object):
    """A class to read the excel file with survival data, get CT/PET path, and prepare train, val and test set

    Args:
        object ([type]): [description]
    """

    def __init__(self, base_path=None,excel_path="FLIP.xlsx", csv_path=None):
        self.base_path = base_path
        self.excel_path=excel_path
        self.csv_path = csv_path
        self.seed = 42  # random state
        self.test_size = 0.0
        self.val_size = 0.2

    def get_data_survival(self, create_csv):
        '''
            retrieve the survival data from the excel file and the path of the nifti files from the base path 
            create a csv file containing all the data if create_csv == True
        '''
        y_time=np.array([])
        y_event=np.array([])
        PT_paths=[]
        CT_paths=[]
        Anonym=[] #anonym name which is also the name of the directory in which there are the PT and CT scans 
        data_exists= True #boolean: path of the nifti file found
        data_dict=dict()

        #open file containing survival information
        excel = openpyxl.load_workbook(self.excel_path)
        sheet = excel.active

        #for each patient
        for i in range(2, sheet.max_row):
            data_exists=True
            x = sheet.cell(row=i, column=5)
            if (x.value != None):
                #retrieve the path to the nifti files from the patient : folder with the Anonymisation name from the excel cell(row=i, column=2)
                path = self.base_path+str(sheet.cell(row=i, column=2).value)+'/'
                nifti_path_PT= glob.glob(os.path.join(path, '**/*_nifti_PT.nii'), recursive=True)
                nifti_path_CT= glob.glob(os.path.join(path, '**/*_nifti_CT.nii'), recursive=True)
                if nifti_path_PT and nifti_path_CT:
                    Anonym = np.append(Anonym,sheet.cell(row=i, column=2).value)
                    PT_paths=np.append(PT_paths, [nifti_path_PT[0]])
                    CT_paths=np.append(CT_paths, [nifti_path_CT[0]])
                else:
                    data_exists=False

                if data_exists:
                    #retrieve y_time and y_event from excel (date format (month/day/year))
                    x = x.value.split('/')
                    x = [int(i) for i in x]
                    diagnosis_date= date(x[2],x[0],x[1])
                    #if censored and there is a last check up date : retrieve date
                    if (sheet.cell(row=i, column=7).value==0 and sheet.cell(row=i, column=9).value!=None):
                        y= sheet.cell(row=i, column=9).value.split('/')
                        y = [int(j) for j in y]
                        last_checkup_date= date(y[2],y[0],y[1])
                    #if not censored and there is a relapse date: retrieve date 
                    elif (sheet.cell(row=i, column=7).value==1 and sheet.cell(row=i, column=8).value!=None):
                        y= sheet.cell(row=i, column=8).value.split('/')
                        y = [int(j) for j in y]
                        last_checkup_date= date(y[2],y[0],y[1])
                    
                    #time is given in 30 days intervals 
                    time= int(((last_checkup_date-diagnosis_date).days)/30)
                    y_time=np.append(y_time, [time])
                    y_event=np.append(y_event,[int(sheet.cell(row=i, column=7).value)])

        y_event= y_event.astype(np.int32)
        y_time=y_time.astype(np.int32)

        if create_csv: 
            #creation of a csv file row: "ID,time,event,CT_path,PT_path"
            data_zip = list(zip(Anonym,y_time, y_event, CT_paths, PT_paths))
            with open(self.csv_path, 'w') as csv_file:
                wtr = csv.writer(csv_file, delimiter=',', lineterminator='\n')
                wtr.writerow(["ID,time,event,CT_path,PT_path"])
                for x in data_zip : wtr.writerow ([str(x[0])+','+str(x[1])+','+str(x[2])+','+str(x[3])+','+str(x[4])])

        return list(zip(PT_paths, CT_paths)), list(zip(y_time, y_event))


    @staticmethod
    def split_train_val_test_split(x, y, test_size=0.0, val_size=0.2, random_state=42):
        x_train, x_val, y_train, y_val = train_test_split(x, y, test_size=val_size, random_state=random_state)
        if test_size!=0.:
            size = test_size/(1 - val_size)
            x_train, x_test, y_train, y_test = train_test_split(x_train, y_train, test_size=size, random_state=random_state)
            return x_train, x_val, x_test, y_train, y_val, y_test
        else:
            return x_train, x_val, y_train, y_val

    def get_images_paths_train_val(self, x_train,x_val):
        dataset_x = dict()
        dataset_x['train']=[]
        dataset_x['val']=[]

        for i in range(len(x_train)):
            dataset_x['train'].append({'pet_img':x_train[i][0], 'ct_img':x_train[i][1]})
        for i in range(len(x_val)):    
            dataset_x['val'].append({'pet_img':x_val[i][0], 'ct_img':x_val[i][1]})
        
        return dataset_x['train'], dataset_x['val']
        






